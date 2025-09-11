from datetime import datetime, timedelta
import sys
import os

# ──[설정]────────────────────────────────────────────────────────
DEPLOY_DATE = datetime(2025, 9, 1)   # 배포일
VALID_DAYS  = 40                     # 사용 가능 기간 (일)
FORCE_EXPIRE = False                 # 테스트 강제 만료 스위치
# ────────────────────────────────────────────────────────────────

def _should_expire(now: datetime) -> bool:
    if FORCE_EXPIRE:
        return True
    if os.environ.get("TEST_EXPIRE", "").strip() == "1":
        return True
    if any(arg in ("--expire-now", "/expire-now") for arg in sys.argv[1:]):
        return True
    expire_date = DEPLOY_DATE + timedelta(days=VALID_DAYS)
    return now > expire_date

def _block_with_message(msg: str):
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk(); root.withdraw()
        messagebox.showerror("사용 기간 만료", msg)
    except Exception:
        print(msg)
    finally:
        sys.exit(1)

if _should_expire(datetime.now()):
    _block_with_message("⚠️ 사용 기간이 만료되었습니다.\n개발자에게 문의하세요.")

# -*- coding: utf-8 -*-
"""
desktop_app.py  (v2025-09-01, tag-based rounds)
- PandaLive 하트 집계 & 쪽지 발송 (Tkinter)
- 엑셀 총합산: 파일명 4자리 태그(예: 0804) 기준 회차/요약 생성
"""

import os, sys, io, re, csv, json, time, threading, subprocess, zipfile, unicodedata
from pathlib import Path
from datetime import datetime
from typing import List, Tuple

import pandas as pd
import numpy as np

import tkinter as tk
from tkinter import (
    Tk, ttk, StringVar, Text, NORMAL, DISABLED, END,
    filedialog, messagebox
)

# ====== helpers ======
def sanitize(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', "_", str(name))[:31] or "Sheet"

def visual_len(val) -> int:
    s = str("" if val is None else val)
    w = 0
    for ch in s:
        if unicodedata.east_asian_width(ch) in ("F","W","A") or ord(ch) >= 0x1F300:
            w += 2
        else:
            w += 1
    return w

def _strip_zw(s: str) -> str:
    return re.sub(r"[\u200b-\u200f\u202a-\u202e\u2060-\u206f\ufeff]", "", s)

BRACKET_ANY = r"[()\[\]{}<>「」『』【】〈〉《》⟦⟧❲❳]"

def normalize_nick(nick: str) -> str:
    if not isinstance(nick, str):
        return ""
    s = unicodedata.normalize("NFKC", nick)
    s = _strip_zw(s)
    for _ in range(3):
        s = re.sub(fr"^\s*{BRACKET_ANY}.*?{BRACKET_ANY}\s*", "", s)
        s = re.sub(fr"\s*{BRACKET_ANY}.*?{BRACKET_ANY}\s*$", "", s)
        s = re.sub(fr"{BRACKET_ANY}.*?{BRACKET_ANY}", " ", s)
    s = re.sub(r"[\U0001F000-\U0001FAFF\U00002700-\U000027BF\U00002600-\U000026FF]+", "", s)
    s = re.sub(r"[❤️♡♥︎💗💖💘💝💞💟✨⭐️☀️]+", "", s)
    s = re.sub(r"\s*(님|형|누나|오빠|언니)$", "", s)
    s = re.sub(r"[\"'`]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def normalize_bj(name: str) -> str:
    if not isinstance(name, str):
        return ""
    s = unicodedata.normalize("NFKC", name)
    s = _strip_zw(s)
    s = re.sub(fr"^\s*{BRACKET_ANY}.*?{BRACKET_ANY}\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def extract_date_from_name(name: str) -> str:
    s = name.lower()
    m = re.search(r'(20\d{2})[.\-_](\d{1,2})[.\-_](\d{1,2})', s)
    if m: return f"{int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
    m = re.search(r'(20\d{2})(\d{2})(\d{2})', s)
    if m: return f"{int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
    m = re.search(r'(\d{2})[.\-_](\d{1,2})[.\-_](\d{1,2})', s)
    if m: return f"{2000+int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
    m = re.search(r'(\d{1,2})(\d{2})', s)
    if m and len(m.group(0)) == 4:
        y = datetime.now().year
        return f"{y:04d}-{int(m[1]):02d}-{int(m[2]):02d}"
    return datetime.now().strftime("%Y-%m-%d")

def read_any_table(path_or_file, sheet=None) -> pd.DataFrame:
    p = str(path_or_file)
    if p.lower().endswith(".xlsx"):
        return pd.read_excel(path_or_file, sheet_name=(sheet if str(sheet).strip() else 0))
    with open(path_or_file, "rb") as f:
        raw = f.read()
    for enc in ["utf-8-sig","utf-8","cp949","euc-kr"]:
        try:
            text = raw.decode(enc)
            try:
                dialect = csv.Sniffer().sniff(text[:4000], delimiters=[",","\t",";","|"])
                sep = dialect.delimiter
            except Exception:
                sep = ","
            return pd.read_csv(io.StringIO(text), sep=sep)
        except Exception:
            continue
    raise ValueError("CSV 인코딩/구분자 해석 실패")

# ===== 경로 상수 =====
BASE = Path(__file__).resolve().parent
RECIP_CSV = BASE / "recipients_preview.csv"
MESSAGE_TXT = BASE / "message.txt"
ENV_FILE = BASE / ".env"
STATUS_JSON = BASE / "send_status.json"
SENDER_PY = BASE / "panda_dm_sender.py"
LOG_OUT = BASE / "sender_stdout.log"
LOG_ERR = BASE / "sender_stderr.log"

FULLWIDTH_SPACE = "\u3000"

def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def load_status(path: Path) -> dict:
    if path.exists():
        try: return json.loads(path.read_text(encoding="utf-8"))
        except: pass
    return {"items": [], "meta": {}}

def save_status(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ---------------- 하트 합계 유틸 ----------------
AFFILIATE_GENERAL_SUBSTRS = ["@ka"]   # '@ka' 포함 시 일반하트 예외

def classify_heart(id_str) -> str:
    if id_str is None:
        return "일반하트"
    s = str(id_str).strip()
    s = s.replace("＠", "@").lower()
    if any(sub in s for sub in AFFILIATE_GENERAL_SUBSTRS):
        return "일반하트"
    return "제휴하트" if "@" in s else "일반하트"

def sanitize_name(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', "_", str(name))[:31] or "BJ"

def preprocess_single(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    col_bj    = next((c for c in df.columns if c == "참여BJ"), None)
    col_heart = next((c for c in df.columns if c == "후원하트"), None)
    col_mix   = next((c for c in df.columns if c == "후원 아이디(닉네임)"), None)
    if not (col_bj and col_heart and col_mix):
        raise ValueError("필수 컬럼 누락: 참여BJ / 후원하트 / 후원 아이디(닉네임)")

    df[col_bj] = df[col_bj].astype(str).str.strip()
    df[col_heart] = df[col_heart].astype(str).str.replace(",", "", regex=False)
    df[col_heart] = pd.to_numeric(df[col_heart], errors="coerce").fillna(0).astype(int)
    df[col_mix] = df[col_mix].astype(str).str.strip()

    sp = df[col_mix].str.extract(r'^\s*(?P<ID>[^()]+?)(?:\((?P<NICK>.*)\))?\s*$')
    df["ID"] = sp["ID"].fillna("").str.replace("＠","@",regex=False).str.strip()
    df["닉네임"] = sp["NICK"].fillna("").apply(normalize_nick)

    base = (
        df.groupby([col_bj, "ID", "닉네임"], as_index=False)[col_heart]
          .sum()
          .rename(columns={col_bj:"참여BJ", col_heart:"후원하트"})
    )
    return base

def make_bj_excel_bytes(bj_name: str, sub_df: pd.DataFrame, admin: bool) -> bytes:
    sub = sub_df.copy()
    # 예외패턴까지 반영한 제휴판별
    sub["is_aff"] = sub["ID"].apply(lambda x: classify_heart(x) == "제휴하트")
    gen = sub[~sub["is_aff"]].sort_values("후원하트", ascending=False)[["ID","닉네임","후원하트"]].copy()
    aff = sub[ sub["is_aff"]].sort_values("후원하트", ascending=False)[["ID","닉네임","후원하트"]].copy()
    gsum, asum = int(gen["후원하트"].sum()), int(aff["후원하트"].sum())

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = sanitize_name(bj_name)
    if admin:
        ws.append(["", bj_name, gsum+asum, "", ""])
        ws.append(["ID","닉네임","후원하트","구분","합계"])
    else:
        ws.append(["", bj_name, gsum+asum])
        ws.append(["ID","닉네임","후원하트"])
    if not gen.empty:
        if admin:
            gen2 = gen.copy(); gen2["구분"] = ""; gen2["합계"] = ""
            rows = gen2.values.tolist(); rows[0][3] = "일반하트"; rows[0][4] = gsum
        else:
            rows = gen.values.tolist()
        for r in rows: ws.append(r)
    if not aff.empty:
        if admin:
            aff2 = aff.copy(); aff2["구분"] = ""; aff2["합계"] = ""
            rows = aff2.values.tolist(); rows[0][3] = "제휴하트"; rows[0][4] = asum
        else:
            rows = aff.values.tolist()
        for r in rows: ws.append(r)

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        m = 0
        for cell in col: m = max(m, visual_len(cell.value))
        ws.column_dimensions[letter].width = max(12, min(m+2, 80))
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

def pack_zip(files: dict[str, bytes]) -> bytes:
    zbio = io.BytesIO()
    with zipfile.ZipFile(zbio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files.items():
            zf.writestr(fname, data)
    zbio.seek(0); return zbio.getvalue()

# ---------------- DM 발송 (탭2) 유틸 ----------------
def guess_columns(df: pd.DataFrame) -> Tuple[str,str,str]:
    cols = [str(c).strip() for c in df.columns]
    id_cands    = ["후원아이디","아이디","ID","id","userId","후원 아이디","후원 아이디(닉네임)"]
    nick_cands  = ["닉네임","후원닉네임","닉","별명","name","nick"]
    heart_cands = ["후원하트","하트","hearts","heart","총하트","하트수"]
    def pick(cands):
        for c in cols:
            if c.replace(" ","") in [x.replace(" ","") for x in cands]:
                return c
        return ""
    id_col = pick(id_cands) or cols[0]
    nick_col = pick(nick_cands) or ""
    heart_col = pick(heart_cands) or cols[-1]
    return id_col, nick_col, heart_col

def normalize_id_from_mix(x: str) -> str:
    if pd.isna(x): return ""
    s = str(x).strip()
    m = re.match(r"^\s*([^()]+)", s)
    return (m.group(1) if m else s).strip()

def normalize_nick_from_mix(x: str) -> str:
    if pd.isna(x): return ""
    s = str(x).strip()
    m = re.search(r"\((.*?)\)", s)
    return (m.group(1).strip() if m else "")

def detect_mixed_id(series: pd.Series, sample: int = 200, threshold: float = 0.3) -> bool:
    try:
        vals = series.dropna().astype(str).head(sample)
        hit = sum(("(" in v and ")" in v and v.find("(") < v.find(")")) for v in vals)
        return (len(vals) > 0) and (hit / len(vals) >= threshold)
    except Exception:
        return False

def prepare_from_csv(df: pd.DataFrame, id_col: str, nick_col: str, heart_col: str, force_mixed: bool):
    tmp = df.copy()
    tmp.columns = [str(c).strip() for c in tmp.columns]
    def _to_int(x):
        s = str(x).strip().replace(",", "")
        try: return int(float(s))
        except: return 0
    series_id = tmp[id_col]
    mixed = force_mixed or detect_mixed_id(series_id)
    if mixed:
        tmp["후원아이디"] = series_id.map(normalize_id_from_mix)
        tmp["닉네임_from_mix"] = series_id.map(normalize_nick_from_mix)
    else:
        tmp["후원아이디"] = series_id.astype(str).str.strip()
        tmp["닉네임_from_mix"] = ""
    tmp["닉네임_src"] = tmp[nick_col].astype(str).str.strip() if nick_col else ""
    tmp["닉네임"] = tmp["닉네임_from_mix"]
    mask_empty = (tmp["닉네임"].astype(str).str.len() == 0)
    tmp.loc[mask_empty, "닉네임"] = tmp.loc[mask_empty, "닉네임_src"]
    tmp["후원하트"] = tmp[heart_col].apply(_to_int)
    agg = (
        tmp.groupby(["후원아이디"], as_index=False)
           .agg(닉네임=("닉네임","first"), 후원하트=("후원하트","sum"))
    )
    auto_df = agg[(agg["후원하트"] >= 1000) & (agg["후원하트"] < 10000)].copy()
    vip_df  = agg[ agg["후원하트"] >= 10000].copy()
    auto_df = auto_df.sort_values(["후원하트","후원아이디"], ascending=[False,True]).reset_index(drop=True)
    vip_df  = vip_df.sort_values(["후원하트","후원아이디"],  ascending=[False,True]).reset_index(drop=True)
    return auto_df, vip_df

def build_messages_with_endspaces(base_msg: str, n: int) -> List[str]:
    FULLWIDTH_SPACE = "\u3000"
    lines = base_msg.splitlines() or [base_msg]
    L = max(1, len(lines))
    out: List[str] = []
    for i in range(n):
        g = i // 5
        add_line_idx = g % L
        add_spaces = (g // L) + 1
        mutated = []
        for j, ln in enumerate(lines):
            mutated.append(ln + (FULLWIDTH_SPACE*add_spaces) if j==add_line_idx else ln)
        msg = "\n".join(mutated)
        out.append(msg[:500] if len(msg)>500 else msg)
    return out

def save_local_bundle(out_df: pd.DataFrame, base_message: str, panda_id: str, panda_pw: str):
    RECIP_CSV.write_text(out_df.to_csv(index=False), encoding="utf-8")
    MESSAGE_TXT.write_text(base_message, encoding="utf-8")
    if panda_id and panda_pw:
        ENV_FILE.write_text(f"PANDA_ID={panda_id}\nPANDA_PW={panda_pw}\n", encoding="utf-8")

# ---------------- GUI ----------------
class App:
    def copy_vip_to_clipboard(self):
        if self._vip_df_cache.empty:
            messagebox.showwarning("안내", "VIP 대상이 없습니다.")
            return
        text = "\n".join(self._vip_df_cache["후원아이디"].astype(str).tolist())
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("완료", f"VIP ID {len(self._vip_df_cache)}명이 클립보드에 복사되었습니다.")

    def export_vip_excel(self):
        if self._vip_df_cache.empty:
            messagebox.showwarning("안내", "VIP 대상이 없습니다.")
            return
        out = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="vip_list.xlsx")
        if not out:
            return
        try:
            cols = [c for c in ["후원아이디","닉네임","후원하트"] if c in self._vip_df_cache.columns]
            self._vip_df_cache[cols].to_excel(out, index=False)
            messagebox.showinfo("완료", f"VIP 엑셀 저장: {out}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 실패: {e}")

    def __init__(self, root: Tk):
        self.root = root
        root.title("하트 합계 & 쪽지 발송 (Desktop)  v2025-09-01")
        root.geometry("1200x820")

        nb = ttk.Notebook(root)
        self.tab_sum = ttk.Frame(nb)
        self.tab_dm  = ttk.Frame(nb)
        nb.add(self.tab_sum, text="📊 하트 합계")
        nb.add(self.tab_dm,  text="✉️ 쪽지 발송")
        nb.pack(fill="both", expand=True)

        self.build_tab_sum()
        self.build_tab_dm()
        self.tick()

    # ----- 탭1 -----
    def build_tab_sum(self):
        f = self.tab_sum
        frm1 = ttk.LabelFrame(f, text="단일 파일 (관리자용/BJ용 ZIP)")
        frm1.pack(fill="x", padx=10, pady=10)
        self.single_path = StringVar(value="(선택 없음)")
        ttk.Label(frm1, textvariable=self.single_path).pack(anchor="w", padx=10, pady=4)
        ttk.Button(frm1, text="파일 선택 (CSV/XLSX)", command=self.pick_single).pack(side="left", padx=10, pady=8)
        ttk.Button(frm1, text="관리자용 ZIP 저장", command=self.save_admin_zip).pack(side="left", padx=5)
        ttk.Button(frm1, text="BJ용 ZIP 저장", command=self.save_bj_zip).pack(side="left", padx=5)

        frm2 = ttk.LabelFrame(f, text="여러 파일 총합산 엑셀")
        frm2.pack(fill="x", padx=10, pady=10)
        self.multi_paths: List[Path] = []
        self.multi_label = StringVar(value="(선택 없음)")
        ttk.Label(frm2, textvariable=self.multi_label).pack(anchor="w", padx=10, pady=4)
        ttk.Button(frm2, text="파일 여러 개 선택", command=self.pick_multi).pack(side="left", padx=10, pady=8)
        ttk.Button(frm2, text="총합산 엑셀 저장", command=self.save_master_excel).pack(side="left", padx=5)

        self.sum_log = Text(f, height=16)
        self.sum_log.pack(fill="both", expand=True, padx=10, pady=10)
        self.log_sum("[안내] 단일 파일은 '참여BJ / 후원하트 / 후원 아이디(닉네임)' 컬럼이 필요합니다.")
        self._single_df = None
        self._admin_zip_bytes = None
        self._bj_zip_bytes = None

    def log_sum(self, msg: str):
        self.sum_log.configure(state=NORMAL); self.sum_log.insert(END, msg.rstrip()+"\n")
        self.sum_log.configure(state=DISABLED); self.sum_log.see(END)

    def pick_single(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/XLSX","*.csv *.xlsx")])
        if not path: return
        self.single_path.set(path)
        try:
            df = read_any_table(Path(path), sheet=None)
            base = preprocess_single(df)
            summary = base.groupby("참여BJ", as_index=False)["후원하트"].sum().sort_values("후원하트", ascending=False)

            admin_files, bj_files = {"요약.xlsx": self._to_excel_bytes(summary)}, {"요약.xlsx": self._to_excel_bytes(summary)}
            for bj in summary["참여BJ"]:
                sub = base[base["참여BJ"] == bj][["ID","닉네임","후원하트"]]
                admin_files[f"{sanitize_name(str(bj))}.xlsx"] = make_bj_excel_bytes(str(bj), sub, admin=True)
                bj_files[f"{sanitize_name(str(bj))}.xlsx"] = make_bj_excel_bytes(str(bj), sub, admin=False)
            self._admin_zip_bytes = pack_zip(admin_files)
            self._bj_zip_bytes = pack_zip(bj_files)
            self._single_df = base
            self.log_sum(f"[완료] 대상 {len(summary)}명 요약 계산/ZIP 작성 완료.")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _to_excel_bytes(self, df: pd.DataFrame) -> bytes:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "요약"
        ws.append(list(df.columns))
        for _, row in df.iterrows(): ws.append(list(row.values))
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            m = 0
            for cell in col: m = max(m, visual_len(cell.value))
            ws.column_dimensions[letter].width = max(12, min(m+2, 80))
        bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

    def save_admin_zip(self):
        if not self._admin_zip_bytes:
            messagebox.showwarning("안내", "먼저 단일 파일을 선택해 주세요."); return
        out = filedialog.asksaveasfilename(defaultextension=".zip", initialfile="BJ별_관리자용.zip")
        if out: Path(out).write_bytes(self._admin_zip_bytes); self.log_sum(f"[저장] {out}")

    def save_bj_zip(self):
        if not self._bj_zip_bytes:
            messagebox.showwarning("안내", "먼저 단일 파일을 선택해 주세요."); return
        out = filedialog.asksaveasfilename(defaultextension=".zip", initialfile="BJ별_BJ용.zip")
        if out: Path(out).write_bytes(self._bj_zip_bytes); self.log_sum(f"[저장] {out}")

    def pick_multi(self):
        paths = filedialog.askopenfilenames(filetypes=[("CSV/XLSX","*.csv *.xlsx")])
        if not paths: return
        self.multi_paths = [Path(p) for p in paths]
        names = "; ".join([Path(p).name for p in paths])
        self.multi_label.set((names[:120] + ("..." if len(names)>120 else "")))
        self.log_sum(f"[선택] 파일 {len(paths)}개")

    def save_master_excel(self, *_ev):
        import time
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        # === 회차 인덱스 매핑 준비 ===
        round_info = []
        seen = set()
        for p in self.multi_paths:
            fn = Path(p).name
            ds = extract_date_from_name(fn)  # YYYY-MM-DD
            if ds in seen:
                continue
            seen.add(ds)

            m4 = re.search(r'(\d{4})', fn)
            if m4:
                tag = m4.group(1)  # 예: '0804'
            else:
                tag = ds[5:].replace('-', '')  # MMDD

            try:
                sort_key = int(tag)
            except:
                sort_key = int(ds.replace('-', ''))  # YYYYMMDD fallback
                tag = ds[5:].replace('-', '')

            round_info.append((sort_key, ds, tag))

        round_info.sort(key=lambda x: x[0])
        date_to_round = {ds: i+1 for i, (_, ds, _) in enumerate(round_info)}
        date_to_tag   = {ds: tag for _, ds, tag in round_info}

        # 재진입 가드/디바운스
        if getattr(self, "_saving_master", False):
            return
        now = time.time()
        if now - getattr(self, "_last_save_master_ts", 0) < 0.8:
            return
        self._saving_master = True

        def _cleanup():
            self._saving_master = False
            self._last_save_master_ts = time.time()

        def _auto_width(ws):
            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                m = 0
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    m = max(m, visual_len(v))
                ws.column_dimensions[letter].width = max(12, min(m + 2, 80))

        def _unique_sheet_name(base: str, used: set[str]) -> str:
            base = sanitize(str(base))[:31] or "Sheet"
            name = base
            i = 2
            while name in used:
                suf = f" ({i})"
                name = base[:31 - len(suf)] + suf
                i += 1
            used.add(name)
            return name

        try:
            if not getattr(self, "multi_paths", None):
                messagebox.showwarning("안내", "먼저 '파일 여러 개 선택'으로 CSV/XLSX 파일을 선택하세요.")
                _cleanup(); return

            # 1) 파일 읽기 + 전처리 + 파일명 태그 부여
            self._seen_tags = []  # 최초 등장 순서 체크용(정렬은 아래에서 숫자 오름차순)
            all_rows, err_files = [], []
            for p in self.multi_paths:
                p = Path(p)
                try:
                    # 파일명에서 4자리 태그 추출(없으면 MMDD 보정)
                    m4 = re.search(r'(\d{4})', p.name)
                    if m4:
                        tag = m4.group(1)  # 예: '0804'
                    else:
                        tag = extract_date_from_name(p.name)[5:].replace('-', '')  # 'MMDD'

                    if tag not in self._seen_tags:
                        self._seen_tags.append(tag)

                    df_in = read_any_table(p, sheet=None)

                    mix_col = "후원 아이디(닉네임)"
                    if mix_col not in df_in.columns:
                        raise ValueError(f"{p.name}: '{mix_col}' 컬럼이 없습니다.")

                    sp = df_in[mix_col].astype(str).str.extract(
                        r'^\s*(?P<ID>[^()]+?)(?:\((?P<NICK>.*)\))?\s*$'
                    )
                    df_in["ID"] = sp["ID"].fillna("").str.replace("＠", "@", regex=False).str.strip()
                    df_in["닉네임"] = sp["NICK"].fillna("").apply(normalize_nick)
                    if "참여BJ" in df_in.columns:
                        df_in["참여BJ"] = df_in["참여BJ"].astype(str).apply(normalize_bj)

                    # 제휴/일반 구분 (예외 포함)
                    df_in["구분"] = df_in["ID"].apply(classify_heart)

                    # 하트 정수화
                    if "후원하트" in df_in.columns:
                        df_in["후원하트"] = (
                            df_in["후원하트"].astype(str).str.replace(",", "", regex=False)
                            .pipe(pd.to_numeric, errors="coerce").fillna(0).astype(int)
                        )

                    # 파일명 태그
                    df_in["회차태그"] = tag

                    cols = ["회차태그", "후원시간", "참여BJ", "ID", "닉네임", "후원하트", "구분"]
                    exist_cols = [c for c in cols if c in df_in.columns]
                    all_rows.append(df_in[exist_cols].copy())

                except Exception as e:
                    err_files.append(f"{p.name}: {e}")

            if err_files:
                self.log_sum("[경고] 일부 파일을 건너뜀:\n  - " + "\n  - ".join(err_files))
            if not all_rows:
                messagebox.showerror("오류", "처리 가능한 파일이 없습니다.")
                _cleanup(); return

            merged = pd.concat(all_rows, ignore_index=True)

            # 안전망 정규화
            if "닉네임" in merged.columns:
                merged["닉네임"] = merged["닉네임"].apply(normalize_nick)
            if "참여BJ" in merged.columns:
                merged["참여BJ"] = merged["참여BJ"].apply(normalize_bj)

            # 2) 회차번호 매핑 (태그 숫자 오름차순)
            tags_sorted = sorted(getattr(self, "_seen_tags", []), key=lambda x: int(x))
            tag_to_round = {tag: i + 1 for i, tag in enumerate(tags_sorted)}

            # 3) 요약_일별 (파일명 태그 기준)
            need = {"회차태그", "참여BJ", "구분", "후원하트"}
            if not need.issubset(set(merged.columns)):
                messagebox.showerror("오류", "필수 컬럼(회차태그/참여BJ/구분/후원하트) 부족으로 요약을 만들 수 없습니다.")
                _cleanup(); return

            piv = (
                merged.groupby(["회차태그", "참여BJ", "구분"], as_index=False)["후원하트"].sum()
                      .pivot(index=["회차태그", "참여BJ"], columns="구분", values="후원하트")
                      .fillna(0)
                      .reset_index()
            )
            for col in ["일반하트", "제휴하트"]:
                if col not in piv.columns:
                    piv[col] = 0
            piv["총합"] = piv["일반하트"] + piv["제휴하트"]
            piv["회차"] = piv["회차태그"].map(tag_to_round).fillna(0).astype(int)
            piv = piv[piv["회차"] > 0].sort_values(["회차", "참여BJ"]).reset_index(drop=True)

            df_daily = piv[["회차", "회차태그", "참여BJ", "일반하트", "제휴하트", "총합"]].rename(columns={"회차태그": "태그"})
            df_daily["회차"] = df_daily["회차"].astype(str) + "회차"

            # 4) 요약_참여BJ_총계
            merged["참여BJ_정규화"] = merged["참여BJ"].apply(normalize_bj)
            total_by_bj = (
                merged.groupby(["참여BJ_정규화", "구분"], as_index=False)["후원하트"].sum()
                      .pivot(index="참여BJ_정규화", columns="구분", values="후원하트")
                      .fillna(0)
                      .reset_index()
                      .rename(columns={"참여BJ_정규화": "참여BJ"})
            )
            for col in ["일반하트", "제휴하트"]:
                if col not in total_by_bj.columns:
                    total_by_bj[col] = 0
            total_by_bj["총합"] = total_by_bj["일반하트"] + total_by_bj["제휴하트"]
            df_total = total_by_bj[["참여BJ", "일반하트", "제휴하트", "총합"]].copy()

            # 5) 저장 경로
            # 첫 파일명에서 날짜 문자열만 활용해 기본 파일명 구성 (실제 집계는 태그 기준)
            default_name = f"총합산_{extract_date_from_name(Path(self.multi_paths[0]).name)}.xlsx"
            out = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name)
            if not out:
                _cleanup(); return

            # 6) 엑셀 작성
            wb = Workbook()

            # (A) 요약_일별
            ws_daily = wb.active; ws_daily.title = "요약_일별"
            ws_daily.append(list(df_daily.columns))
            for row in df_daily.itertuples(index=False):
                ws_daily.append(list(row))
            _auto_width(ws_daily)

            # (B) 요약_참여BJ_총계
            ws_total = wb.create_sheet(title="요약_참여BJ_총계")
            ws_total.append(list(df_total.columns))
            for row in df_total.sort_values("총합", ascending=False).itertuples(index=False):
                ws_total.append(list(row))
            per_round = pd.DataFrame(columns=["회차번호", "후원하트", "회차태그"])  # 안전한 기본값
            if {"회차태그", "후원하트"}.issubset(merged.columns):
                per_round = (
                    merged.groupby("회차태그", as_index=False)["후원하트"].sum()
                )
                per_round["회차번호"] = per_round["회차태그"].map(tag_to_round)
                per_round = per_round.dropna(subset=["회차번호"]).sort_values("회차번호")
                per_round["회차번호"] = per_round["회차번호"].astype(int)

            if not per_round.empty:
                ws_total.append([])
                ws_total.append(["회차별 전체 합계"])
                ws_total.append(["회차번호", "후원하트", "회차태그"])
                for r in per_round.to_dict("records"):
                    ws_total.append([r["회차번호"], int(r["후원하트"]), r["회차태그"]])
    
            _auto_width(ws_total)

            # (C) 참여BJ별 상세 + 회차별 합계(태그 기준)
            merged_sorted = merged.copy()
            sort_cols = [c for c in ["회차태그", "후원시간"] if c in merged_sorted.columns]
            if sort_cols:
                merged_sorted = merged_sorted.sort_values(sort_cols)
            merged_sorted["BJ_KEY"] = merged_sorted["참여BJ"].apply(normalize_bj)

            used_names = {"요약_일별", "요약_참여BJ_총계"}
            for bj_key, sub in merged_sorted.groupby("BJ_KEY", dropna=False):
                bj_key = bj_key if isinstance(bj_key, str) and bj_key.strip() else "미지정BJ"

                gsum = int(sub.loc[sub["구분"] == "일반하트", "후원하트"].sum())
                asum = int(sub.loc[sub["구분"] == "제휴하트", "후원하트"].sum())
                tsum = gsum + asum

                sheet_title = _unique_sheet_name(bj_key, used_names)
                ws = wb.create_sheet(title=sheet_title)

                ws.append([f"총 일반하트={gsum}", f"총 제휴하트={asum}", f"총합={tsum}"])

                cols = ["회차태그", "후원시간", "ID", "닉네임", "후원하트", "구분"]
                exist_cols = [c for c in cols if c in sub.columns]
                ws.append(exist_cols)
                for row in sub[exist_cols].itertuples(index=False):
                    ws.append(list(row))

                # 하단: 회차별 합계 (파일명 태그 기준)
                if "회차태그" in sub.columns:
                    per_round = sub.groupby("회차태그", as_index=False)["후원하트"].sum()
                    per_round["회차번호"] = per_round["회차태그"].map(tag_to_round).fillna(0).astype(int)
                    per_round = per_round[per_round["회차번호"] > 0].sort_values("회차번호")
                    if not per_round.empty:
                        ws.append([])
                        ws.append(["회차별 합계"])
                        ws.append(["회차", "하트합계", "회차태그"])
                        for _, r in per_round.iterrows():
                            ws.append([f"{int(r['회차번호'])}회차",int(r["후원하트"]),str(r["회차태그"])])

                _auto_width(ws)

            wb.save(out)
            self.log_sum(f"[저장] 총합산 엑셀 저장: {out}")
            messagebox.showinfo("완료", f"총합산 엑셀 저장 완료:\n{out}")

        except PermissionError:
            messagebox.showerror("저장 실패", "엑셀에서 해당 파일이 열려있습니다.\n파일을 닫고 다시 저장하세요.")
        except Exception as e:
            messagebox.showerror("저장 실패", f"엑셀 저장 중 오류가 발생했습니다:\n{e}")
        finally:
            _cleanup()

    # ----- 탭2 (쪽지 전송: 기존 그대로) -----
    def build_tab_dm(self):
        f = self.tab_dm

        frm_top = ttk.Frame(f); frm_top.pack(fill="x", padx=10, pady=8)
        left = ttk.Frame(frm_top); right = ttk.Frame(frm_top)
        left.pack(side="left", fill="both", expand=True)
        right.pack(side="left", fill="both", expand=True, padx=(10,0))

        ttk.Label(left, text="팬더 아이디").pack(anchor="w")
        self.var_pid = StringVar(value=""); ttk.Entry(left, textvariable=self.var_pid).pack(fill="x", pady=2)
        ttk.Label(left, text="팬더 비밀번호").pack(anchor="w")
        self.var_ppw = StringVar(value=""); ttk.Entry(left, textvariable=self.var_ppw, show="*").pack(fill="x", pady=2)
        ttk.Label(left, text="기본 쪽지 (여러 줄)").pack(anchor="w", pady=(6,0))
        self.txt_msg = Text(left, height=8); self.txt_msg.pack(fill="both", expand=True, pady=2)

        ttk.Label(right, text="원본 CSV 업로드").pack(anchor="w")
        ttk.Button(right, text="CSV 선택", command=self.pick_recip_csv).pack(anchor="w", pady=2)
        ttk.Label(right, text="수동 ID 입력(줄바꿈/쉼표/공백)").pack(anchor="w", pady=(6,0))
        self.txt_manual = Text(right, height=6); self.txt_manual.pack(fill="both", expand=True, pady=2)

        frm_mid = ttk.Frame(f); frm_mid.pack(fill="x", padx=10, pady=8)
        ttk.Button(frm_mid, text="💾 파일 저장(.env/CSV/MSG)", command=self.save_bundle).pack(side="left", padx=2)
        ttk.Button(frm_mid, text="메시지 변형 미리보기", command=self.preview_messages).pack(side="left", padx=2)

        self.lbl_counts = StringVar(value="자동발송 대상: 0명 | VIP: 0명")
        ttk.Label(f, textvariable=self.lbl_counts).pack(anchor="w", padx=12, pady=(0,4))

        frm_lists = ttk.Frame(f); frm_lists.pack(fill="x", padx=10, pady=4)
        left_list = ttk.LabelFrame(frm_lists, text="자동발송 대상 (1,000~9,999)")
        right_list = ttk.LabelFrame(frm_lists, text="VIP 대상 (10,000+)")
        left_list.pack(side="left", fill="both", expand=True, padx=(0,5))
        right_list.pack(side="left", fill="both", expand=True, padx=(5,0))

        self.tree_auto = ttk.Treeview(left_list, columns=("id","nick","heart"), show="headings", height=6)
        for c,t in zip(("id","nick","heart"),("후원아이디","닉네임","하트")):
            self.tree_auto.heading(c, text=t)
        self.tree_auto.pack(fill="both", expand=True)

        self.tree_vip = ttk.Treeview(right_list, columns=("id","nick","heart"), show="headings", height=6)
        for c,t in zip(("id","nick","heart"),("후원아이디","닉네임","하트")):
            self.tree_vip.heading(c, text=t)
        self.tree_vip.pack(fill="both", expand=True)

        vip_btns = ttk.Frame(right_list)
        vip_btns.pack(fill="x", padx=4, pady=4)
        ttk.Button(vip_btns, text="VIP ID 복사", command=self.copy_vip_to_clipboard).pack(side="left", padx=2)
        ttk.Button(vip_btns, text="VIP 엑셀 저장", command=self.export_vip_excel).pack(side="left", padx=2)

        frm_run = ttk.Frame(f); frm_run.pack(fill="x", padx=10, pady=8)
        ttk.Label(frm_run, text="시작 인덱스").pack(side="left", padx=3)
        self.var_start = StringVar(value="0"); ttk.Entry(frm_run, textvariable=self.var_start, width=6).pack(side="left")
        ttk.Label(frm_run, text="최대 인원(0=전원)").pack(side="left", padx=3)
        self.var_limit = StringVar(value="0"); ttk.Entry(frm_run, textvariable=self.var_limit, width=6).pack(side="left")
        self.headless = StringVar(value="1"); ttk.Checkbutton(frm_run, text="헤드리스 실행", variable=self.headless, onvalue="1", offvalue="0").pack(side="left", padx=10)
        self.reset_status = StringVar(value="0"); ttk.Checkbutton(frm_run, text="현황 초기화", variable=self.reset_status, onvalue="1", offvalue="0").pack(side="left", padx=6)
        ttk.Button(frm_run, text="📨 전송 실행", command=self.start_sender).pack(side="left", padx=10)
        ttk.Button(frm_run, text="⛔ 강제 종료", command=self.kill_sender).pack(side="right", padx=4)
        ttk.Button(frm_run, text="🧹 현황/임시 파일 삭제", command=self.cleanup_files).pack(side="right", padx=4)

        frm_dash = ttk.LabelFrame(f, text="실시간 현황 / 로그"); frm_dash.pack(fill="both", expand=True, padx=10, pady=8)
        self.lbl_stats = StringVar(value="총 대상: 0 | 성공: 0 | 실패: 0 | 대기: 0")
        ttk.Label(frm_dash, textvariable=self.lbl_stats).pack(anchor="w", padx=8, pady=4)
        ttk.Label(frm_dash, text="상태").pack(anchor="w", padx=8)
        self.log_out = Text(frm_dash, height=10); self.log_out.pack(fill="both", expand=True, padx=8)
        ttk.Label(frm_dash, text="무시").pack(anchor="w", padx=8)
        self.log_err = Text(frm_dash, height=6); self.log_err.pack(fill="both", expand=True, padx=8)

        self.sender_pid = None
        self._auto_df_cache = pd.DataFrame(columns=["후원아이디","닉네임","후원하트"])
        self._vip_df_cache  = pd.DataFrame(columns=["후원아이디","닉네임","후원하트"])

    def _fill_tree(self, tree: ttk.Treeview, rows: list[tuple]):
        for i in tree.get_children(): tree.delete(i)
        for r in rows: tree.insert("", "end", values=r)

    def pick_recip_csv(self):
        path = filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
        if not path: return
        try:
            df = pd.read_csv(path)
        except Exception:
            df = pd.read_csv(path, encoding="utf-8-sig")

        id_col, nick_col, heart_col = guess_columns(df)
        mixed_guess = detect_mixed_id(df[id_col])
        auto_df, vip_df = prepare_from_csv(df, id_col, nick_col, heart_col, force_mixed=mixed_guess)
        self._auto_df_cache = auto_df
        self._vip_df_cache  = vip_df

        self._fill_tree(self.tree_auto, [(r["후원아이디"], r.get("닉네임",""), r["후원하트"]) for _,r in auto_df.head(50).iterrows()])
        self._fill_tree(self.tree_vip,  [(r["후원아이디"], r.get("닉네임",""), r["후원하트"]) for _,r in vip_df.head(50).iterrows()])

        self.lbl_counts.set(f"자동발송 대상: {len(auto_df)}명 | VIP: {len(vip_df)}명")
        messagebox.showinfo("완료", f"자동발송 {len(auto_df)}명 / VIP {len(vip_df)}명 추출 완료 (미리보기 상위 50명 표시).")

    def save_bundle(self):
        manual = self.txt_manual.get("1.0", END).strip()
        if manual:
            tokens = [t.strip() for t in re.split(r"[,\s]+", manual) if t.strip()]
            tokens = list(dict.fromkeys(tokens))
            out_df = pd.DataFrame({"후원아이디": tokens, "닉네임": ["" for _ in tokens], "후원하트": [1000 for _ in tokens]})
        else:
            out_df = self._auto_df_cache

        base_message = self.txt_msg.get("1.0", END).rstrip("\n")
        panda_id = self.var_pid.get().strip()
        panda_pw = self.var_ppw.get().strip()

        if out_df.empty:
            messagebox.showwarning("안내", "대상자 목록이 비었습니다."); return

        save_local_bundle(out_df, base_message, panda_id, panda_pw)
        messagebox.showinfo("완료", f"recipients_preview.csv / message.txt / .env 저장 완료\n(자동발송 대상 {len(out_df)}명)")
        st_json = {"items":[{"index":int(i),"id":str(r["후원아이디"]),"status":"pending","updated":now_ts()} for i,r in out_df.iterrows()],
                   "meta":{"created": now_ts()}}
        save_status(STATUS_JSON, st_json)

    def preview_messages(self):
        manual = self.txt_manual.get("1.0", END).strip()
        if manual:
            tokens = [t.strip() for t in re.split(r"[,\s]+", manual) if t.strip()]
            tokens = list(dict.fromkeys(tokens))
            out_df = pd.DataFrame({"후원아이디": tokens, "닉네임": ["" for _ in tokens], "후원하트": [1000 for _ in tokens]})
        else:
            out_df = self._auto_df_cache

        base_message = self.txt_msg.get("1.0", END).rstrip("\n")
        msgs = build_messages_with_endspaces(base_message, len(out_df))
        sample = "\n\n".join(f"[{i}] {r['후원아이디']}\n{msgs[i]}" for i,(_,r) in enumerate(out_df.head(5).iterrows()))
        messagebox.showinfo("미리보기 (상위 5명)", sample or "없음")

    def start_sender(self):
        if not RECIP_CSV.exists() or not MESSAGE_TXT.exists():
            messagebox.showwarning("안내", "먼저 파일 저장(.env/CSV/MSG)을 누르세요.")
            return

        try:
            s = int(self.var_start.get().strip() or "0")
            l = int(self.var_limit.get().strip() or "0")
        except:
            messagebox.showerror("오류", "시작/최대 인원은 정수")
            return

        headless = (self.headless.get() == "1")
        reset = (self.reset_status.get() == "1")

        try:
            LOG_OUT.write_text("", encoding="utf-8")
            LOG_ERR.write_text("", encoding="utf-8")
        except:
            pass

        def _run_inside():
            try:
                import panda_dm_sender as sender
                sender.run_from_gui(
                    headless=headless,
                    status_file=str(STATUS_JSON),
                    reset=reset,
                    start=s,
                    limit=l,
                )
            except Exception as e:
                try:
                    with open(LOG_ERR, "a", encoding="utf-8") as f:
                        f.write(f"{now_ts()}  {e}\n")
                except:
                    pass
                messagebox.showerror("실행 오류", f"전송 실행 중 오류:\n{e}")

        threading.Thread(target=_run_inside, daemon=True).start()
        messagebox.showinfo("안내", "전송을 시작했습니다. 현황/로그를 확인하세요.")

    def kill_sender(self):
        if self.sender_pid:
            try:
                if os.name == "nt": subprocess.run(["taskkill","/PID", str(self.sender_pid), "/F", "/T"])
                else: os.kill(self.sender_pid, 9)
                self.sender_pid = None; messagebox.showinfo("완료", "프로세스 종료")
            except Exception as e:
                messagebox.showerror("오류", f"종료 실패: {e}")
        else:
            messagebox.showinfo("안내", "실행 중인 전송 프로세스가 없습니다.")

    def cleanup_files(self):
        for p in [STATUS_JSON, RECIP_CSV, MESSAGE_TXT, ENV_FILE]:
            try: p.unlink(missing_ok=True)
            except: pass
        try: LOG_OUT.unlink(missing_ok=True); LOG_ERR.unlink(missing_ok=True)
        except: pass
        messagebox.showinfo("완료", "현황/임시 파일 삭제 완료")

    def refresh_dashboard(self):
        data = load_status(STATUS_JSON)
        items = data.get("items", [])
        total = len(items)
        succ = sum(1 for x in items if x.get("status")=="success")
        fail = sum(1 for x in items if x.get("status")=="fail")
        pend = total - succ - fail
        self.lbl_stats.set(f"총 대상: {total} | 성공: {succ} 🟢 | 실패: {fail} 🔴 | 대기: {pend} 🟡")
        lines = []
        for row in items[-30:]:
            s = row.get("status","pending")
            lamp = "🟢" if s=="success" else ("🔴" if s=="fail" else "🟡")
            lines.append(f"{lamp}  #{row.get('index')}  {row.get('id')}  {row.get('updated')}")
        table_text = "\n".join(lines) if lines else "(진행 항목 없음)"
        def tail_bytes(p: Path, max_chars=12000):
            if not p.exists(): return ""
            try: return p.read_text(encoding="utf-8")[-max_chars:]
            except: return ""
        self._set_text(self.log_out, table_text + "\n\n--- STDOUT ---\n" + tail_bytes(LOG_OUT))
        self._set_text(self.log_err, tail_bytes(LOG_ERR))

    def _set_text(self, widget: Text, content: str):
        widget.configure(state=NORMAL); widget.delete("1.0", END); widget.insert("1.0", content or "(없음)")
        widget.configure(state=DISABLED); widget.see(END)

    def tick(self):
        try: self.refresh_dashboard()
        except Exception: pass
        self.root.after(1000, self.tick)

# ===== main =====
if __name__ == "__main__":
    root = Tk()
    app = App(root)
    root.mainloop()
