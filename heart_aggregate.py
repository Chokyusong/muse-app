# heart_aggregate.py
import io, re, csv, zipfile, unicodedata
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===== 여기부터는 기존 app.py “내용”을 함수로 감쌈 =====
def show():
    # ↓↓↓ 당신이 이전에 준 app.py 로직을 그대로 옮기되,
    #     st.set_page_config(...) 한 줄만 빼고 모두 포함시키면 됩니다.
    #     (아래는 핵심만 그대로 가져다 둔 버전)

    st.subheader("BJ별 하트 정리 자동화")
    st.caption("단일 파일 → 관리자용/BJ용 ZIP (합산), 여러 파일 → 총합산 엑셀 (요약 + 참여BJ별 시트)")

    # -------------------- helpers --------------------
    def visual_len(val) -> int:
        s = str(val) if val is not None else ""
        w = 0
        for ch in s:
            if unicodedata.east_asian_width(ch) in ("F", "W", "A"):
                w += 2
            elif ord(ch) >= 0x1F300:
                w += 2
            else:
                w += 1
        return w

    def autosize_columns(wb, min_w=12, max_w=80, pad=2):
        for ws in wb.worksheets:
            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                max_width = 0
                for cell in col:
                    if cell.value is not None:
                        max_width = max(max_width, visual_len(cell.value))
                ws.column_dimensions[letter].width = max(min_w, min(max_width + pad, max_w))

    def sanitize(name: str) -> str:
        return re.sub(r'[\\/*?:\[\]]', "_", str(name))[:31] or "BJ"

    def normalize_nick(nick: str) -> str:
        if not isinstance(nick, str):
            return ""
        nick = re.sub(r'^\[.*?\]', '', nick)
        nick = re.sub(r'\(.*?\)', '', nick)
        return nick.strip()

    def normalize_bj(name: str) -> str:
        if not isinstance(name, str):
            return ""
        return re.sub(r'^\[.*?\]', '', name).strip()

    @st.cache_data(show_spinner=False, persist=False, ttl=0, max_entries=10)
    def read_any_table(uploaded_file, sheet: str | int | None):
        name = (uploaded_file.name or "").lower()
        if name.endswith(".xlsx"):
            return pd.read_excel(uploaded_file, sheet_name=(sheet if str(sheet).strip() else 0))
        raw = uploaded_file.read(); uploaded_file.seek(0)
        for enc in ["utf-8", "utf-8-sig", "cp949", "euc-kr"]:
            try:
                text = raw.decode(enc)
                try:
                    dialect = csv.Sniffer().sniff(text[:4000], delimiters=[",", "\t", ";", "|"])
                    sep = dialect.delimiter
                except Exception:
                    sep = ","
                return pd.read_csv(io.StringIO(text), sep=sep)
            except Exception:
                continue
        raise ValueError("CSV 인코딩/구분자 해석 실패")

    # ---------------- 단일 파일 (관리자용/BJ용 ZIP) ----------------
    def preprocess(df: pd.DataFrame) -> pd.DataFrame:
        df.columns = [str(c).strip() for c in df.columns]
        col_bj    = next((c for c in df.columns if c == "참여BJ"), None)
        col_heart = next((c for c in df.columns if c == "후원하트"), None)
        col_mix   = next((c for c in df.columns if c == "후원 아이디(닉네임)"), None)
        if not (col_bj and col_heart and col_mix):
            raise ValueError("필수 컬럼 누락: 참여BJ / 후원하트 / 후원 아이디(닉네임)")

        df[col_bj] = df[col_bj].astype(str).str.strip()
        df[col_heart] = df[col_heart].astype(str).str.replace(",", "", regex=False)
        df[col_heart] = pd.to_numeric(df[col_heart], errors="coerce").fillna(0).astype(int)
        df[col_mix] = df[col_mix].astype(str).str.strip()

        sp = df[col_mix].str.extract(r'^\s*(?P<ID>[^()]+?)(?:\((?P<NICK>.*)\))?\s*$')
        df["ID"] = sp["ID"].fillna("").str.replace("＠","@",regex=False).str.strip()
        df["닉네임"] = sp["NICK"].fillna("").apply(normalize_nick)

        base = (
            df.groupby([col_bj, "ID", "닉네임"], as_index=False)[col_heart]
              .sum()
              .rename(columns={col_bj:"참여BJ", col_heart:"후원하트"})
        )
        return base

    def _xlsx_bytes_from_df(writer_fn) -> bytes:
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as w:
            writer_fn(w)
        bio.seek(0)
        wb = load_workbook(bio)
        autosize_columns(wb)
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return out.getvalue()

    def make_bj_excel(bj_name: str, sub_df: pd.DataFrame, admin: bool) -> bytes:
        sub = sub_df.copy()
        sub["is_aff"] = sub["ID"].str.contains("@")
        gen = sub[~sub["is_aff"]].sort_values("후원하트", ascending=False)[["ID","닉네임","후원하트"]].copy()
        aff = sub[ sub["is_aff"]].sort_values("후원하트", ascending=False)[["ID","닉네임","후원하트"]].copy()
        gsum, asum = int(gen["후원하트"].sum()), int(aff["후원하트"].sum())
        total = gsum + asum
        sheet = sanitize(bj_name)

        def _write(w):
            if admin:
                row1 = pd.DataFrame([[ "", bj_name, total, "", "" ]],
                                    columns=["ID","닉네임","후원하트","구분","합계"])
            else:
                row1 = pd.DataFrame([[ "", bj_name, total ]],
                                    columns=["ID","닉네임","후원하트"])
            row1.to_excel(w, sheet_name=sheet, index=False, header=False, startrow=0)

            if admin:
                pd.DataFrame(columns=["ID","닉네임","후원하트","구분","합계"]).to_excel(
                    w, sheet_name=sheet, index=False, startrow=1)
            else:
                pd.DataFrame(columns=["ID","닉네임","후원하트"]).to_excel(
                    w, sheet_name=sheet, index=False, startrow=1)

            row = 2
            if not gen.empty:
                blk = gen.copy()
                if admin:
                    blk["구분"], blk["합계"] = "", ""
                    blk.iloc[0, blk.columns.get_loc("구분")] = "일반하트"
                    blk.iloc[0, blk.columns.get_loc("합계")] = gsum
                blk.to_excel(w, sheet_name=sheet, index=False, header=False, startrow=row)
                row += len(blk)
            if not aff.empty:
                blk = aff.copy()
                if admin:
                    blk["구분"], blk["합계"] = "", ""
                    blk.iloc[0, blk.columns.get_loc("구분")] = "제휴하트"
                    blk.iloc[0, blk.columns.get_loc("합계")] = asum
                blk.to_excel(w, sheet_name=sheet, index=False, header=False, startrow=row)

        return _xlsx_bytes_from_df(_write)

    def build_file_sets(base: pd.DataFrame):
        summary = base.groupby("참여BJ", as_index=False)["후원하트"].sum().sort_values("후원하트", ascending=False)

        def make_summary_bytes() -> bytes:
            return _xlsx_bytes_from_df(lambda w: summary.to_excel(w, sheet_name="요약", index=False))

        def pack_zip(files: dict[str, bytes]) -> bytes:
            zbio = io.BytesIO()
            with zipfile.ZipFile(zbio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for fname, data in files.items():
                    zf.writestr(fname, data)
            zbio.seek(0); return zbio.getvalue()

        admin_files, bj_files = {"요약.xlsx": make_summary_bytes()}, {"요약.xlsx": make_summary_bytes()}
        for bj in summary["참여BJ"]:
            sub = base[base["참여BJ"] == bj][["ID","닉네임","후원하트"]]
            admin_files[f"{sanitize(bj)}.xlsx"] = make_bj_excel(str(bj), sub, admin=True)
            bj_files[f"{sanitize(bj)}.xlsx"] = make_bj_excel(str(bj), sub, admin=False)
        return (admin_files, pack_zip(admin_files)), (bj_files, pack_zip(bj_files))

    # ---------------- 여러 파일 (총합산 엑셀) ----------------
    def extract_date_from_name(name: str) -> str:
        s = name.lower()
        m = re.search(r'(20\d{2})[.\-_](\d{1,2})[.\-_](\d{1,2})', s)
        if m: return f"{int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
        m = re.search(r'(20\d{2})(\d{2})(\d{2})', s)
        if m: return f"{int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
        m = re.search(r'(\d{2})[.\-_](\d{1,2})[.\-_](\d{1,2})', s)
        if m: return f"{2000+int(m[1]):04d}-{int(m[2]):02d}-{int(m[3]):02d}"
        m = re.search(r'(\d{1,2})(\d{2})', s)
        if m and len(m.group(0)) == 4:
            y = datetime.now().year
            return f"{y:04d}-{int(m[1]):02d}-{int(m[2]):02d}"
        return datetime.now().strftime("%Y-%m-%d")

    def build_master_excel_bytes(merged_df, df_daily, df_total) -> bytes:
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as w:
            df_daily.to_excel(w, index=False, sheet_name="요약_일별")
            df_total.to_excel(w, index=False, sheet_name="요약_참여BJ_총계")
            merged_df = merged_df.copy()
            merged_df["참여BJ_정규화"] = merged_df["참여BJ"].apply(normalize_bj)
            sort_cols = [c for c in ["날짜","후원시간"] if c in merged_df.columns]
            if sort_cols:
                merged_df = merged_df.sort_values(sort_cols)
            for bj, sub in merged_df.groupby("참여BJ_정규화"):
                gsum = int(sub.loc[sub["구분"]=="일반하트","후원하트"].sum())
                asum = int(sub.loc[sub["구분"]=="제휴하트","후원하트"].sum())
                tsum = gsum + asum
                top = pd.DataFrame([[f"총 일반하트={gsum}", f"총 제휴하트={asum}", f"총합={tsum}"]])
                cols = ["날짜","후원시간","ID","닉네임","후원하트","구분"]
                exist_cols = [c for c in cols if c in sub.columns]
                out = sub[exist_cols].reset_index(drop=True)
                sheet = sanitize(bj)
                top.to_excel(w, index=False, header=False, sheet_name=sheet, startrow=0)
                out.to_excel(w, index=False, sheet_name=sheet, startrow=2)
        bio.seek(0); wb = load_workbook(bio); autosize_columns(wb)
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return out.getvalue()

    # ================== UI ==================
    uploaded = st.file_uploader("단일 CSV/XLSX 업로드 (관리자용/BJ용 ZIP 생성 — 합산)", type=["csv", "xlsx"])
    sheet_name = st.text_input("시트 이름 (엑셀일 때만)", value="")

    if uploaded:
        try:
            df_in = read_any_table(uploaded, sheet_name if uploaded.name.lower().endswith(".xlsx") else None)
            base = preprocess(df_in)
            (admin_files, admin_zip), (bj_files, bj_zip) = build_file_sets(base)
            left, right = st.columns(2, gap="large")
            with left:
                st.subheader("관리자용 (합산, 구분/합계 포함)")
                st.download_button("📦 관리자용 ZIP 다운로드", data=admin_zip,
                                   file_name="BJ별_관리자용.zip", mime="application/zip",
                                   use_container_width=True, key="zip-admin")
            with right:
                st.subheader("BJ용 (합산, 심플버전)")
                st.download_button("📦 BJ용 ZIP 다운로드", data=bj_zip,
                                   file_name="BJ별_BJ용.zip", mime="application/zip",
                                   use_container_width=True, key="zip-bj")
        except Exception as e:
            st.error(f"오류: {e}")

    st.header("여러 파일 합산 (총합산 엑셀 생성)")
    multi = st.file_uploader("여러 CSV/XLSX 업로드", type=["csv","xlsx"], accept_multiple_files=True)

    if multi:
        all_rows = []
        for uf in multi:
            try:
                date_str = extract_date_from_name(uf.name)
                df_in = read_any_table(uf, None)
                mix_col = "후원 아이디(닉네임)"
                if mix_col not in df_in.columns:
                    raise ValueError(f"{uf.name}: '{mix_col}' 컬럼이 없습니다.")
                sp = df_in[mix_col].astype(str).str.extract(r'^\s*(?P<ID>[^()]+?)(?:\((?P<NICK>.*)\))?\s*$')
                df_in["ID"] = sp["ID"].fillna("").str.replace("＠","@",regex=False).str.strip()
                df_in["닉네임"] = sp["NICK"].fillna("").apply(normalize_nick)
                df_in["구분"] = np.where(df_in["ID"].str.contains("@"), "제휴하트", "일반하트")
                df_in["날짜"] = date_str
                cols = ["날짜","후원시간","참여BJ","ID","닉네임","후원하트","구분"]
                exist_cols = [c for c in cols if c in df_in.columns]
                all_rows.append(df_in[exist_cols])
            except Exception as e:
                st.warning(f"{uf.name} 처리 오류: {e}")

        if all_rows:
            merged = pd.concat(all_rows, ignore_index=True)
            need_cols = {"날짜","참여BJ","구분","후원하트"}
            if not need_cols.issubset(set(merged.columns)):
                st.error("필수 컬럼(날짜/참여BJ/구분/후원하트) 부족으로 요약을 만들 수 없습니다.")
            else:
                piv = (merged.groupby(["날짜","참여BJ","구분"], as_index=False)["후원하트"].sum()
                             .pivot(index=["날짜","참여BJ"], columns="구분", values="후원하트")
                             .fillna(0).reset_index())
                for col in ["일반하트","제휴하트"]:
                    if col not in piv.columns: piv[col] = 0
                piv["총합"] = piv["일반하트"] + piv["제휴하트"]
                daily_out = piv[["날짜","참여BJ","일반하트","제휴하트","총합"]].sort_values(["날짜","참여BJ"]).reset_index(drop=True)
                st.subheader("요약_일별"); st.dataframe(daily_out, use_container_width=True, hide_index=True)

                merged["참여BJ_정규화"] = merged["참여BJ"].apply(normalize_bj)
                total_by_bj = (merged.groupby(["참여BJ_정규화","구분"], as_index=False)["후원하트"].sum()
                                      .pivot(index="참여BJ_정규화", columns="구분", values="후원하트")
                                      .fillna(0).reset_index()
                                      .rename(columns={"참여BJ_정규화":"참여BJ"}))
                for col in ["일반하트","제휴하트"]:
                    if col not in total_by_bj.columns: total_by_bj[col] = 0
                total_by_bj["총합"] = total_by_bj["일반하트"] + total_by_bj["제휴하트"]
                st.subheader("요약_참여BJ_총계 (정규화 적용)")
                st.dataframe(total_by_bj.sort_values("총합", ascending=False), use_container_width=True, hide_index=True)

                master_bytes = build_master_excel_bytes(
                    merged_df=merged,
                    df_daily=daily_out,
                    df_total=total_by_bj[["참여BJ","일반하트","제휴하트","총합"]]
                )
                st.download_button("📥 총합산 엑셀 다운로드",
                                   data=master_bytes,
                                   file_name="총합산.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
